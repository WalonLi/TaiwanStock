
import twstock
import time
import datetime
import os
import re
import openpyxl
import sys

# type,code,name,ISIN,start,market,group,CFI
# 股票,1101,台泥,TW0001101004,1962/02/09,上市,水泥工業,ESVUFR
#
# DATATUPLE = namedtuple('Data', ['date', 'capacity', 'turnover', 'open',
#                                 'high', 'low', 'close', 'change', 'transaction'])




class Global:
    wait_time = 6
    skip = True # some stock information can't get.

def handle_history_data(stock, y, m, raw_path, sheet):
    Global.skip = True

    for loop in range(3):
        try:
            if twstock.codes[stock.sid].market == '上市':
                time.sleep(Global.wait_time) # small delay for website block
                Global.wait_time -= 1
                if Global.wait_time < 2:
                    Global.wait_time = 6
            else:
                time.sleep(1)

            date_list = stock.fetch(y, m)

            sheet.append([str(y) + " %02d"%m,
                         min(stock.low),
                         max(stock.high),
                         date_list[0].open,
                         date_list[-1].close,
                         sum(stock.capacity),
                         round(date_list[-1].close - date_list[0].open, 2),
                         sum(stock.turnover)])
            # flush raw data
            file = open(raw_path, 'w')
            for i in range(len(date_list)):
                file.write(str(date_list[i])+'\n')
            file.close()
            Global.skip = False
            break
        except ValueError as e:
            print('unknown value error.', e)
            
def handle_data(stock, y, m, raw_path, sheet):
    for loop in range(3):
        try:
            if twstock.codes[stock.sid].market == '上市':
                time.sleep(Global.wait_time) # small delay for website block
                Global.wait_time -= 1
                if Global.wait_time < 2:
                    Global.wait_time = 6
            else:
                time.sleep(1)

            date_list = stock.fetch(y, m)
            
            # find location to update.
            match = False
            for row in sheet.rows:
                tag = str(y) + " %02d"%m
                if row[0].value == tag:
                    sheet[str(row[1].column) + str(row[0].row)].value = min(stock.low)
                    sheet[str(row[2].column) + str(row[0].row)].value = max(stock.high)
                    sheet[str(row[3].column) + str(row[0].row)].value = date_list[0].open
                    sheet[str(row[4].column) + str(row[0].row)].value = date_list[-1].close
                    sheet[str(row[5].column) + str(row[0].row)].value = sum(stock.capacity)
                    sheet[str(row[6].column) + str(row[0].row)].value = round(date_list[-1].close - date_list[0].open, 2)
                    sheet[str(row[7].column) + str(row[0].row)].value = sum(stock.turnover)
                    match = True
                    break
            if not match:
                sheet.insert_rows(1, 1)
                sheet['A2'].value = str(y) + " %02d"%m
                sheet['B2'].value = min(stock.low)
                sheet['C2'].value = max(stock.high)
                sheet['D2'].value = date_list[0].open
                sheet['E2'].value = date_list[-1].close
                sheet['F2'].value = sum(stock.capacity)
                sheet['G2'].value = round(date_list[-1].close - date_list[0].open, 2)
                sheet['H2'].value = sum(stock.turnover)

            file = open(raw_path, 'w')
            for i in range(len(date_list)):
                file.write(str(date_list[i]) + '\n')
            file.close()
            Global.skip = False
            break
        except ValueError as e:
            print('unknown value error2.', e)


history_head = ['日期', '最低', '最高', '開盤', '收盤', '成交量', '價差', '金額']

def get_history():
    # for number in range(3031, 4000, 1):
    for number in range(3000, 4000, 1):

        if str(number) in twstock.codes:

            # time.sleep(2) # small delay for website block

            stock_info = twstock.codes[str(number)]
            stock_name = stock_info[1] + re.sub('\*', '', stock_info[2]) # strip some special character
            if '-DR' in stock_name:
                continue

            path = 'StockList\\' + stock_info[6]
            if not os.path.isdir(path):
                os.mkdir(path)

            path += '\\' + stock_name
            if not os.path.isdir(path):
                os.mkdir(path)

            excel_path = path + '\\' + stock_name + '_History.xlsx'
            if not os.path.isfile(excel_path):
                # if not exist, get all data
                file = openpyxl.Workbook()
                sheet = file.active
                sheet.title = 'Sheet1'
                sheet.append(history_head)
                
                start_date = stock_info[4].split('/')
                start_date = datetime.date(int(start_date[0]), int(start_date[1]), int(start_date[2]))

                target_date = datetime.date(2002, 11, 1)
                if start_date > target_date:
                    target_date = start_date

                year = datetime.date.today().year
                month = datetime.date.today().month

                stock = twstock.Stock(str(number))
                print(stock_name, start_date)
                
                
                i = 0
                try:
                    while True:
                        # handle data
                        print('    %d %d parsing...' % (year, month))
                        if not os.path.isfile(path + '\\raw%d%02d' % (year, month)):
                            handle_history_data(stock, year, month, path + '\\raw%d%02d' % (year, month), sheet)
                            if Global.skip:
                                error_file = open(path+'\\raw%d%02d.fail'%(year,month), 'w')
                                error_file.close()
                                break

                        if target_date.year == year and target_date.month == month:
                            print(target_date.year, target_date.month, ' end')
                            break
                        i += 1
                        month -= 1
                        if month == 0:
                            year -= 1
                            month = 12
                    file.save(excel_path)
                    file.close()
                except:
                    file.save(excel_path)
                    # file.close()
                    print('get_history: unknown connect error ...')
                    time.sleep(60)

            
            
def fix_history():
    for root, dirs, files in os.walk(os.getcwd() + '\\StockList'):
        files.reverse()

        for f in files:
            if f[-5:] == '.fail':
                file = openpyxl.load_workbook(root + '\\' + files[-1])
                sheet = file.get_sheet_by_name('Sheet1')
                
                print(root, files[-1], f)
                number = int(files[-1][:4])
                fail_year = int(f[3:7])
                fail_month = int(f[7:9])
                print('Stock:%d Y:%d M:%02d' % (number, fail_year, fail_month))
                start_date = twstock.codes[str(number)][4].split('/')
                start_date = datetime.date(int(start_date[0]), int(start_date[1]), int(start_date[2]))

                target_date = datetime.date(2002, 11, 1)
                if start_date > target_date:
                    target_date = start_date

                year = fail_year
                month = fail_month

                stock = twstock.Stock(str(number))
                
                i = 0
                try:
                    while True:
                        # handle data
                        print('    %d %d parsing...' % (year, month))
                        if not os.path.isfile(root +'\\raw%d%02d'%(year,month)):
                            handle_history_data(stock, year, month, root + '\\raw%d%02d' % (year, month), sheet)
                            if Global.skip:
                                error_file = open(root +'\\raw%d%02d.fail'%(year,month), 'w')
                                error_file.close()
                                if year == datetime.date.today().year and month == datetime.date.today().month:
                                    break
                            else:
                                if os.path.isfile(root + '\\' + f):
                                    os.remove(root + '\\' + f)

                        if target_date.year == year and target_date.month == month:
                            print(target_date.year, target_date.month, ' end')
                            break
                        i += 1
                        month -= 1
                        if month == 0:
                            year -= 1
                            month = 12
                    file.save(root + '\\' + files[-1])
                    file.close()
                except:
                    file.save(root + '\\' + files[-1])
                    file.close()
                    print('fix_history: unknown connect error ...')
                    time.sleep(60)
                break # only scan once

def get_all_this_month(year, month):
    for root, dirs, files in os.walk(os.getcwd() + '\\StockList'):
        for f in files:
            if f[-5:] == '.xlsx':
                try:
                    number = int(f[:4])
                    stock = twstock.Stock(str(number))
                    # if twstock.codes[str(number)].market == '上市':
                    #     time.sleep(Global.wait_time) # small delay for website block
                    #     Global.wait_time -= 1
                    #     if Global.wait_time < 2:
                    #         Global.wait_time = 6
                        
                    print(root, f)
                    file = openpyxl.load_workbook(root + '\\' + f)
                    sheet = file.get_sheet_by_name('Sheet1')
    
                    handle_data(stock, year, month, root + '\\raw%d%02d' % (year, month), sheet)
                    if Global.skip:
                        error_file = open(root + '\\raw%d%02d.fail' % (year, month), 'w')
                        error_file.close()
                    else:
                        if os.path.isfile(root + '\\' + '\\raw%d%02d.fail'):
                            os.remove(root + '\\' + '\\raw%d%02d.fail')
                except:
                    print('get_history: unknown connect error3 ...')
                    time.sleep(60)

                file.save(root + '\\' + f)
                file.close()
     
    
if __name__ == '__main__':

    print('Choose functions.')
    print('11:Generate History.')
    print('22:Fix unresolved History.')
    print('33:Get data with specific month.')
    
    option = input()
    if option == '11':
        get_history()
    elif option == '22':
        fix_history()
    elif option == '33':
        year = input('year:')
        month = input('month:')
        get_all_this_month(int(year), int(month))
